"""Render an .xlsx workbook to a self-contained HTML document.

Used by the ingestion preview UI so reviewers can see the source spreadsheet
roughly as it appears in Excel — preserving merged cells, header rows, column
widths, bold text, and basic fills. Not pixel-perfect; good enough for review.
"""

from __future__ import annotations

import html
import io
from datetime import date, datetime, time
from typing import Iterable

import openpyxl
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


_WEEKDAY_NAMES = (
    "Sunday", "Monday", "Tuesday", "Wednesday",
    "Thursday", "Friday", "Saturday",
)
_MONTH_NAMES = (
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
)


def _format_value(value, number_format: str | None) -> str:
    """Render a cell value using its Excel number format hints."""
    if value is None:
        return ""
    fmt = (number_format or "").lower()
    if isinstance(value, datetime) or isinstance(value, date):
        d = value if isinstance(value, datetime) else datetime(value.year, value.month, value.day)
        # Custom format hints — handle the common ones first.
        if "dddd" in fmt:
            return _WEEKDAY_NAMES[(d.weekday() + 1) % 7]
        if "ddd" in fmt:
            return _WEEKDAY_NAMES[(d.weekday() + 1) % 7][:3]
        if "mmmm" in fmt:
            return _MONTH_NAMES[d.month - 1]
        if "mmm" in fmt and "d" not in fmt and "y" not in fmt:
            return _MONTH_NAMES[d.month - 1][:3]
        # Year is 1900 → almost always a weekday-formatted cell mis-read; fall
        # back to weekday name rather than printing "1900-01-02".
        if d.year == 1900:
            return _WEEKDAY_NAMES[(d.weekday() + 1) % 7]
        if "yyyy" in fmt and "m" not in fmt and "d" not in fmt:
            return str(d.year)
        return d.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, float):
        # Trim trailing .0 on whole numbers; otherwise keep two-decimal precision
        # if the original format suggests it.
        if value.is_integer():
            return str(int(value))
        return f"{value:g}"
    return str(value)


def _cell_style(cell: Cell) -> str:
    bits: list[str] = []
    font = cell.font
    if font is not None:
        if font.bold:
            bits.append("font-weight:600")
        if font.color and font.color.rgb and isinstance(font.color.rgb, str) and len(font.color.rgb) == 8:
            # ARGB → RGB
            bits.append(f"color:#{font.color.rgb[2:].lower()}")
    fill = cell.fill
    if fill is not None and fill.fgColor and getattr(fill.fgColor, "type", None) == "rgb":
        rgb = fill.fgColor.rgb
        if isinstance(rgb, str) and len(rgb) == 8 and rgb != "00000000":
            bits.append(f"background-color:#{rgb[2:].lower()}")
    align = cell.alignment
    if align is not None:
        if align.horizontal in {"left", "right", "center"}:
            bits.append(f"text-align:{align.horizontal}")
        if align.vertical in {"top", "middle", "bottom"}:
            v = "middle" if align.vertical == "center" else align.vertical
            bits.append(f"vertical-align:{v}")
    return ";".join(bits)


def _merged_lookup(ws: Worksheet) -> tuple[dict[tuple[int, int], tuple[int, int]], set[tuple[int, int]]]:
    """Return (anchor → (rowspan, colspan)) and a set of cells covered by another merge."""
    anchors: dict[tuple[int, int], tuple[int, int]] = {}
    covered: set[tuple[int, int]] = set()
    for rng in ws.merged_cells.ranges:
        rowspan = rng.max_row - rng.min_row + 1
        colspan = rng.max_col - rng.min_col + 1
        anchors[(rng.min_row, rng.min_col)] = (rowspan, colspan)
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                if (r, c) != (rng.min_row, rng.min_col):
                    covered.add((r, c))
    return anchors, covered


def _column_widths(ws: Worksheet, max_col: int) -> list[int]:
    """Approximate column widths in pixels (Excel: width × 7px). Defaults to 80px."""
    widths: list[int] = []
    for c in range(1, max_col + 1):
        letter = get_column_letter(c)
        dim = ws.column_dimensions.get(letter)
        if dim and dim.width:
            widths.append(max(40, int(dim.width * 7)))
        else:
            widths.append(80)
    return widths


def _dense_bounding_box(ws: Worksheet) -> tuple[int, int]:
    """Find the bottom-right corner of the dense block anchored at A1.

    Excel files often have stray cells far from the main table (data-validation
    sources, leftover scratch cells, scratch padding rows). The reviewer only
    cares about the dense block in the upper-left.

    Strategy: walk left→right one column at a time. A column "belongs" if it
    has any content in the row range we've already accepted *or* if it adds
    new bottom rows that connect to existing content. Once we find a column
    that's both empty within current rows and doesn't extend the table down,
    stop. Same shape applies bottom-up.
    """
    if ws.max_row is None or ws.max_column is None:
        return (0, 0)

    # Find the last meaningful row in column A (or the leftmost few cols) by
    # scanning down until we hit a stretch of empty rows. This is our seed.
    GAP_TOLERANCE = 2  # consecutive empty rows allowed before bailing
    seed_last_row = 1
    empty_streak = 0
    for r in range(1, ws.max_row + 1):
        if any(ws.cell(r, c).value not in (None, "") for c in (1, 2, 3)):
            seed_last_row = r
            empty_streak = 0
        else:
            empty_streak += 1
            if empty_streak > GAP_TOLERANCE and seed_last_row > 1:
                break

    # Now extend the column range using the seed row range.
    last_col = 1
    empty_col_streak = 0
    for c in range(1, ws.max_column + 1):
        col_has_content = any(
            ws.cell(r, c).value not in (None, "") for r in range(1, seed_last_row + 1)
        )
        if col_has_content:
            last_col = c
            empty_col_streak = 0
        else:
            empty_col_streak += 1
            if empty_col_streak > 1 and last_col >= 1:
                break

    return (seed_last_row, last_col)


def _render_sheet(ws: Worksheet, bounded: bool = True) -> str:
    if ws.max_row is None or ws.max_column is None:
        return ""
    if bounded:
        bound_row, bound_col = _dense_bounding_box(ws)
    else:
        bound_row, bound_col = ws.max_row, ws.max_column
    if bound_row == 0 or bound_col == 0:
        return ""
    anchors, covered = _merged_lookup(ws)
    widths = _column_widths(ws, bound_col)
    parts: list[str] = []
    parts.append('<table class="sheet">')
    parts.append("<colgroup>")
    for w in widths:
        parts.append(f'<col style="width:{w}px">')
    parts.append("</colgroup>")
    parts.append("<tbody>")
    for r in range(1, bound_row + 1):
        parts.append("<tr>")
        for c in range(1, bound_col + 1):
            if (r, c) in covered:
                continue
            cell = ws.cell(row=r, column=c)
            value_html = html.escape(_format_value(cell.value, cell.number_format))
            style = _cell_style(cell)
            attrs = []
            if (r, c) in anchors:
                rowspan, colspan = anchors[(r, c)]
                # Clip merged spans to the bounding box.
                rowspan = min(rowspan, bound_row - r + 1)
                colspan = min(colspan, bound_col - c + 1)
                if rowspan > 1:
                    attrs.append(f'rowspan="{rowspan}"')
                if colspan > 1:
                    attrs.append(f'colspan="{colspan}"')
            if style:
                attrs.append(f'style="{style}"')
            parts.append(f'<td {" ".join(attrs)}>{value_html}</td>')
        parts.append("</tr>")
    parts.append("</tbody></table>")
    return "".join(parts)


_BASE_CSS = """
:root { color-scheme: light dark; }
body { margin: 0; padding: 16px; font-family: 'Segoe UI', Calibri, Arial, sans-serif; font-size: 13px; background: transparent; color: #1a1a1a; }
@media (prefers-color-scheme: dark) { body { color: #e6e6e6; } }
.sheet-name { margin: 18px 0 6px; font-size: 13px; font-weight: 600; opacity: 0.7; }
.sheet-name:first-child { margin-top: 0; }
table.sheet { border-collapse: collapse; table-layout: fixed; }
table.sheet td { border: 1px solid rgba(127,127,127,0.35); padding: 4px 8px; vertical-align: top; word-wrap: break-word; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
"""


def render_xlsx_to_html(content: bytes, bounded: bool = True) -> str | None:
    """Render an xlsx byte blob to a standalone HTML document. Returns None on failure.

    When bounded=True (default), clips to the dense top-left rectangle so
    off-screen junk (data-validation source lists, leftover scratch cells)
    doesn't leak into the preview. Pass bounded=False to render the entire
    sheet — useful as an escape hatch in the UI.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        return None
    sheet_html_parts: list[str] = []
    for ws in wb.worksheets:
        rendered = _render_sheet(ws, bounded=bounded)
        if not rendered:
            continue
        if len(wb.worksheets) > 1:
            sheet_html_parts.append(f'<div class="sheet-name">{html.escape(ws.title)}</div>')
        sheet_html_parts.append(rendered)
    if not sheet_html_parts:
        return None
    body = "".join(sheet_html_parts)
    return f"<!doctype html><html><head><meta charset='utf-8'><style>{_BASE_CSS}</style></head><body>{body}</body></html>"


def _xls_dense_bounding_box(sheet) -> tuple[int, int]:
    """xlrd equivalent of _dense_bounding_box. Uses 0-based indexing internally."""
    if sheet.nrows == 0 or sheet.ncols == 0:
        return (0, 0)
    GAP_TOLERANCE = 2
    seed_last_row = 0
    empty_streak = 0
    for r in range(sheet.nrows):
        if any(sheet.cell_value(r, c) not in ("", None) for c in range(min(3, sheet.ncols))):
            seed_last_row = r
            empty_streak = 0
        else:
            empty_streak += 1
            if empty_streak > GAP_TOLERANCE and seed_last_row > 0:
                break
    last_col = 0
    empty_col_streak = 0
    for c in range(sheet.ncols):
        col_has_content = any(
            sheet.cell_value(r, c) not in ("", None) for r in range(seed_last_row + 1)
        )
        if col_has_content:
            last_col = c
            empty_col_streak = 0
        else:
            empty_col_streak += 1
            if empty_col_streak > 1:
                break
    return (seed_last_row + 1, last_col + 1)  # convert back to 1-based count


def _format_xls_cell(workbook, cell) -> str:
    """Format an xlrd cell, handling date types and weekday remap."""
    import xlrd
    from datetime import datetime as _dt
    if cell.ctype == xlrd.XL_CELL_EMPTY or cell.ctype == xlrd.XL_CELL_BLANK:
        return ""
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            t = xlrd.xldate_as_tuple(cell.value, workbook.datemode)
            if t[0] == 1900:
                return _WEEKDAY_NAMES[(_dt(*t[:3]).weekday() + 1) % 7]
            if t[0]:
                return _dt(*t[:3]).strftime("%Y-%m-%d")
            return f"{t[3]:02d}:{t[4]:02d}"
        except Exception:
            return str(cell.value)
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        v = cell.value
        if v == int(v):
            return str(int(v))
        return f"{v:g}"
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return "TRUE" if cell.value else "FALSE"
    return str(cell.value)


def render_xls_to_html(content: bytes, bounded: bool = True) -> str | None:
    """Render a legacy .xls byte blob to HTML via xlrd. Returns None on failure.

    See render_xlsx_to_html for the bounded= flag semantics.
    """
    try:
        import xlrd
        workbook = xlrd.open_workbook(file_contents=content, formatting_info=False)
    except Exception:
        return None
    sheet_html_parts: list[str] = []
    for sheet in workbook.sheets():
        if bounded:
            bound_row, bound_col = _xls_dense_bounding_box(sheet)
        else:
            bound_row, bound_col = sheet.nrows, sheet.ncols
        if bound_row == 0 or bound_col == 0:
            continue
        # merged_cells: list of (rlo, rhi, clo, chi) — half-open ranges, 0-based.
        anchors: dict[tuple[int, int], tuple[int, int]] = {}
        covered: set[tuple[int, int]] = set()
        for rlo, rhi, clo, chi in getattr(sheet, "merged_cells", []) or []:
            rowspan = rhi - rlo
            colspan = chi - clo
            anchors[(rlo + 1, clo + 1)] = (rowspan, colspan)
            for r in range(rlo, rhi):
                for c in range(clo, chi):
                    if (r, c) != (rlo, clo):
                        covered.add((r + 1, c + 1))
        parts = ['<table class="sheet"><tbody>']
        for r in range(bound_row):
            parts.append("<tr>")
            for c in range(bound_col):
                if (r + 1, c + 1) in covered:
                    continue
                value_html = html.escape(_format_xls_cell(workbook, sheet.cell(r, c)))
                attrs = []
                if (r + 1, c + 1) in anchors:
                    rowspan, colspan = anchors[(r + 1, c + 1)]
                    rowspan = min(rowspan, bound_row - r)
                    colspan = min(colspan, bound_col - c)
                    if rowspan > 1:
                        attrs.append(f'rowspan="{rowspan}"')
                    if colspan > 1:
                        attrs.append(f'colspan="{colspan}"')
                parts.append(f'<td {" ".join(attrs)}>{value_html}</td>')
            parts.append("</tr>")
        parts.append("</tbody></table>")
        if len(workbook.sheets()) > 1:
            sheet_html_parts.append(f'<div class="sheet-name">{html.escape(sheet.name)}</div>')
        sheet_html_parts.append("".join(parts))
    if not sheet_html_parts:
        return None
    body = "".join(sheet_html_parts)
    return f"<!doctype html><html><head><meta charset='utf-8'><style>{_BASE_CSS}</style></head><body>{body}</body></html>"


def render_csv_to_html(content: bytes, encoding: str = "utf-8") -> str | None:
    import csv as _csv
    try:
        text = content.decode(encoding, errors="replace")
        rows = list(_csv.reader(io.StringIO(text)))
    except Exception:
        return None
    if not rows:
        return None
    parts = ['<table class="sheet"><tbody>']
    for row in rows:
        parts.append("<tr>" + "".join(f"<td>{html.escape(cell)}</td>" for cell in row) + "</tr>")
    parts.append("</tbody></table>")
    body = "".join(parts)
    return f"<!doctype html><html><head><meta charset='utf-8'><style>{_BASE_CSS}</style></head><body>{body}</body></html>"
